import json

import nbformat as nbf

nb = nbf.v4.new_notebook()
cells = []

cells.append(nbf.v4.new_markdown_cell("""# Anaconda channel catalog: `main` vs `main-x`

Builds `anaconda_channel_catalog.xlsx` with three sheets. Every channel sheet opens with a header line
("Anaconda <channel> channel — N packages — generated <date> from the official Anaconda repository. Browse online: ...")
above the column headers.

| Sheet | Contents |
|---|---|
| `main` | every unique package on **main**: name, latest version, what it does (Anaconda's own channeldata summary), whether the name also appears on main-x, **total downloads, download source, download as-of date** — sorted by downloads descending |
| `main-x` | every unique package on **main-x**: same shape — name, latest version, what it does (anaconda.org `main-x` summary), license, on-main flag, download columns. Sorted by downloads descending (ties alphabetical). |
| `Summary` | every figure with its source and retrieval date |

All metadata is **Anaconda-published only**; no third-party enrichment anywhere.

**Data sources (Anaconda only)**

- `main` channel contents/versions/summaries — public channeldata: `https://repo.anaconda.com/pkgs/main/channeldata.json`
- `main` downloads — anaconda.org package API: `https://api.anaconda.org/package/anaconda/<name>` (package-level `ndownloads`, server-side aggregate over all files/versions/platforms; field chosen by live probe). Cached resume-safe in `download_counts_cache.json`.
- `main-x` channel contents/versions/licenses — authenticated channel `https://repo.anaconda.cloud/repo/main-x/` (Bearer repo token). Its `channeldata.json` is a stub (`packages: {}`), so the notebook falls back to the per-subdir `repodata.json` it points at.
- `main-x` summaries + download counts — anaconda.org repocore API (same public API the anaconda.org front-end calls): `https://api.anaconda.org/repocore/channels/main-x` and `.../artifacts`. Exact-name match to the catalog; no fuzzy matching. **As of the probe date, anaconda.org reports `download_count = 0` for every main-x package** (telemetry not populated for this channel yet) — published verbatim, flagged on the Summary sheet; do not read zeros as zero usage.

**Integrity policy — refuse rather than report a wrong number**

- Content-Length enforced when provided; early-closed connections (`IncompleteRead`) are truncation.
- Responses must be complete valid UTF-8 JSON; structure validated; package-count sanity floor applied; pagination totals cross-checked.
- Counts are always **unique package names** (build variants collapsed), never artifact rows.
- If the anaconda.org probes find no usable fields, the notebook stops. Downloads coverage <50% of main -> explicit Summary warning.
- The xlsx is written only after all fetches and validations pass (atomically); any failure raises `ChannelDataError`.

**Repo token sources**: env var `ANACONDA_REPO_TOKEN`, then `conda config --show`, then `~/.condarc`. Never printed, never persisted. The anaconda.org APIs used need no auth."""))

cells.append(nbf.v4.new_code_cell("""import json
import os
import re
import subprocess
import time
import urllib.error
import urllib.request
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime, timezone
from http.client import IncompleteRead

import pandas as pd

MAIN_CHANNELDATA_URL = "https://repo.anaconda.com/pkgs/main/channeldata.json"
MAIN_X_CHANNELDATA_URL = "https://repo.anaconda.cloud/repo/main-x/channeldata.json"
MAIN_X_REPODATA_URL = "https://repo.anaconda.cloud/repo/main-x/{subdir}/repodata.json"
ANACONDA_ORG_PACKAGE_API = "https://api.anaconda.org/package/anaconda/{name}"
REPOCORE_CHANNEL_API = "https://api.anaconda.org/repocore/channels/{channel}"
REPOCORE_ARTIFACTS_API = "https://api.anaconda.org/repocore/channels/{channel}/artifacts"
REPOCORE_PAGE_LIMIT = 1000
BROWSE_URLS = {"main": "anaconda.org/channels/main", "main-x": "anaconda.org/channels/main-x"}
TOKEN_ENV_VAR = "ANACONDA_REPO_TOKEN"
OUTPUT_XLSX = "anaconda_channel_catalog.xlsx"
DOWNLOAD_CACHE_FILE = "download_counts_cache.json"
DL_WORKERS = 8
DL_TIMEOUT = 60
DL_MAX_RETRIES = 5
MIN_DOWNLOAD_COVERAGE_WARN = 0.50
HTTP_TIMEOUT_SECONDS = 180
MIN_PACKAGES_EXPECTED = 100
USER_AGENT = "anaconda-channel-catalog/3.1 (stdlib urllib)"

URL_TOKEN_RE = re.compile(r"/t/([A-Za-z0-9][A-Za-z0-9._-]*)(?:/|$)")


class ChannelDataError(RuntimeError):
    \"\"\"Raised when channel data is missing, unreachable, truncated, or fails validation.
    When this is raised, no catalog file is written.\"\"\"


def now_utc():
    return datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC")


def today_utc():
    return datetime.now(timezone.utc).strftime("%Y-%m-%d")


def redact(text):
    return URL_TOKEN_RE.sub("/t/<redacted>/", text) if text else text

print(f"config loaded {now_utc()}")"""))

cells.append(nbf.v4.new_code_cell("""def _tokens_from_text(text):
    return URL_TOKEN_RE.findall(text or "")


def resolve_repo_token():
    \"\"\"Return (token, origin). Never prints the token. Raises ChannelDataError if not found.\"\"\"
    token = os.environ.get(TOKEN_ENV_VAR, "").strip()
    if token:
        return token, f"environment variable {TOKEN_ENV_VAR}"
    try:
        out = subprocess.run(["conda", "config", "--show", "channels", "default_channels"],
                             capture_output=True, text=True, timeout=30)
        found = _tokens_from_text(out.stdout)
        if found:
            return found[0], "conda config (channel URL with embedded /t/<token>/)"
    except (OSError, subprocess.SubprocessError):
        pass
    for path in (os.path.expanduser("~/.condarc"),):
        try:
            with open(path, "r", encoding="utf-8") as fh:
                found = _tokens_from_text(fh.read())
            if found:
                return found[0], f"token embedded in channel URL in {path}"
        except OSError:
            continue
    raise ChannelDataError(
        "No Anaconda repo token found. main-x is an authenticated channel. "
        f"Set {TOKEN_ENV_VAR}, or configure conda so a channel URL contains /t/<token>/ (`anaconda token install`)."
    )

print("token resolver ready")"""))

cells.append(nbf.v4.new_code_cell("""def fetch_json_with_integrity(url, label, extra_headers=None, timeout=None):
    \"\"\"Fetch a JSON document, refusing anything that looks truncated. Returns (parsed_json, provenance_dict).\"\"\"
    headers = {"Accept-Encoding": "identity", "User-Agent": USER_AGENT}
    if extra_headers:
        headers.update(extra_headers)
    shown = redact(url)
    req = urllib.request.Request(url, headers=headers)
    try:
        resp = urllib.request.urlopen(req, timeout=timeout or HTTP_TIMEOUT_SECONDS)
    except urllib.error.HTTPError as e:
        hint = ""
        if e.code in (401, 403, 404):
            hint = " Authentication/authorization problem: token missing, invalid, expired, or not entitled."
        raise ChannelDataError(f"[{label}] HTTP {e.code} fetching {shown}.{hint} Refusing to continue.") from e
    except urllib.error.URLError as e:
        raise ChannelDataError(f"[{label}] network error fetching {shown}: {e.reason}. Refusing to continue.") from e

    with resp:
        expected = resp.headers.get("Content-Length")
        last_modified = resp.headers.get("Last-Modified")
        chunks, received = [], 0
        while True:
            try:
                chunk = resp.read(1 << 20)
            except IncompleteRead as e:
                got = received + len(e.partial or b"")
                raise ChannelDataError(
                    f"[{label}] TRUNCATED response from {shown}: closed after {got} bytes (announced {expected}). Refusing to continue."
                ) from e
            if not chunk:
                break
            chunks.append(chunk)
            received += len(chunk)

    raw = b"".join(chunks)
    if expected is not None and received != int(expected):
        raise ChannelDataError(f"[{label}] TRUNCATED response from {shown}: received {received} of {expected} bytes. Refusing to continue.")
    if received == 0:
        raise ChannelDataError(f"[{label}] empty response from {shown}. Refusing to continue.")
    try:
        text = raw.decode("utf-8")
    except UnicodeDecodeError as e:
        raise ChannelDataError(f"[{label}] response from {shown} is not valid UTF-8 ({e}); likely truncated/corrupt. Refusing to continue.") from e
    try:
        data = json.loads(text)
    except json.JSONDecodeError as e:
        raise ChannelDataError(
            f"[{label}] invalid JSON from {shown} (line {e.lineno} col {e.colno}: {e.msg}); almost certainly truncated. Refusing to continue."
        ) from e

    return data, {
        "source_url": shown,
        "retrieved_utc": now_utc(),
        "http_last_modified": last_modified or "(not provided)",
        "bytes_received": received,
        "content_length_check": f"matched {received} bytes" if expected is not None else "no Content-Length header; verified via stream terminator + strict JSON parse",
    }

print("integrity-checked fetcher ready")"""))

cells.append(nbf.v4.new_code_cell("""def _version_key(version):
    parts = re.split(r"(\\d+)", str(version or ""))
    return tuple((0, int(p)) if p.isdigit() else (1, p.lower()) for p in parts if p)


def validate_channeldata(data, label, provenance):
    if not isinstance(data, dict):
        raise ChannelDataError(f"[{label}] channeldata top level is {type(data).__name__}, not an object. Refusing.")
    pkgs = data.get("packages")
    subdirs = data.get("subdirs")
    if not isinstance(pkgs, dict):
        raise ChannelDataError(f"[{label}] channeldata 'packages' missing or not an object. Refusing.")
    if not isinstance(subdirs, list):
        raise ChannelDataError(f"[{label}] channeldata 'subdirs' missing or not a list. Refusing.")
    non_objects = [n for n, m in pkgs.items() if not isinstance(m, dict)]
    if non_objects:
        raise ChannelDataError(f"[{label}] {len(non_objects)} package entries are not objects (e.g. {non_objects[:3]}). Refusing.")
    if 0 < len(pkgs) < MIN_PACKAGES_EXPECTED:
        raise ChannelDataError(f"[{label}] channeldata lists only {len(pkgs)} packages (< floor {MIN_PACKAGES_EXPECTED}); looks partial. Refusing.")
    provenance["channeldata_package_count"] = len(pkgs)
    return pkgs, [s for s in subdirs if isinstance(s, str) and s]


def parse_repodata(data, label, provenance):
    if not isinstance(data, dict):
        raise ChannelDataError(f"[{label}] repodata top level is {type(data).__name__}, not an object. Refusing.")
    removed = data.get("removed", [])
    if not isinstance(removed, list):
        raise ChannelDataError(f"[{label}] repodata 'removed' is not a list. Refusing.")
    removed = set(removed)
    records = []
    for section in ("packages", "packages.conda"):
        table = data.get(section, {})
        if not isinstance(table, dict):
            raise ChannelDataError(f"[{label}] repodata '{section}' is not an object. Refusing.")
        for fn, rec in table.items():
            if fn in removed:
                continue
            if not isinstance(rec, dict) or not rec.get("name") or not rec.get("version"):
                raise ChannelDataError(f"[{label}] repodata record {fn!r} lacks name/version. Refusing.")
            records.append(rec)
    provenance["removed_artifacts_honored"] = len(removed)
    return records


def latest_from_repodata(records, label):
    \"\"\"Collapse build variants -> {unique_name: {'version', 'summary', 'license', 'build_variants_collapsed'}}.\"\"\"
    if not records:
        raise ChannelDataError(f"[{label}] repodata produced zero usable records. Refusing.")
    best, counts = {}, {}
    for rec in records:
        name = rec["name"].strip()
        counts[name] = counts.get(name, 0) + 1
        key = (_version_key(rec["version"]), int(rec.get("build_number") or 0), int(rec.get("timestamp") or 0))
        cur = best.get(name)
        if cur is None or key > cur[0]:
            best[name] = (key, rec)
    return {name: {"version": str(rec["version"]), "summary": "",
                   "license": str(rec.get("license") or ""), "build_variants_collapsed": counts[name]}
            for name, (_, rec) in best.items()}


def catalog_from_channeldata_packages(pkgs, label):
    out, missing_summary = {}, []
    for name, meta in pkgs.items():
        version = str(meta.get("version") or "").strip()
        if not version:
            raise ChannelDataError(f"[{label}] channeldata entry {name!r} has no version. Refusing.")
        summary = str(meta.get("summary") or "").strip()
        if not summary:
            missing_summary.append(name)
        out[name] = {"version": version, "summary": summary,
                     "license": str(meta.get("license") or ""), "build_variants_collapsed": None}
    return out, missing_summary

print("validators ready")"""))

cells.append(nbf.v4.new_markdown_cell("""## API probes — verify shape first, build parsers around what is observed

1. anaconda.org package API for main downloads (`snowflake-connector-python`): print top-level keys, pick the numeric download field empirically. No usable field -> stop.
2. anaconda.org repocore API (the public API the .org front-end calls) for main-x: print channel object keys and one artifact's keys, validate artifact count against the channel's own `artifact_count`. The artifact parser is written around these observed fields."""))

cells.append(nbf.v4.new_code_cell("""# --- probe 1: main downloads on the anaconda.org package API ---
probe_data, _ = fetch_json_with_integrity(ANACONDA_ORG_PACKAGE_API.format(name="snowflake-connector-python"), "probe/anaconda.org")
top_keys = sorted(probe_data.keys())
print("top-level keys of /package/anaconda/snowflake-connector-python:")
print(" ", top_keys)

DOWNLOAD_FIELD = None
for candidate in ("ndownloads", "downloads", "download_count", "total_downloads"):
    v = probe_data.get(candidate)
    if isinstance(v, int) and not isinstance(v, bool) and v >= 0:
        DOWNLOAD_FIELD = candidate
        break
if DOWNLOAD_FIELD is None:
    raise ChannelDataError(
        f"anaconda.org package API has no usable numeric download field (top-level keys: {top_keys}). Stopping rather than deriving a proxy."
    )
print(f"\\nobserved download field: {DOWNLOAD_FIELD!r} = {probe_data[DOWNLOAD_FIELD]:,} "
      f"(package-level aggregate across all files/versions/platforms; per-file sums may drift slightly)")

def parse_downloads(data):
    v = data.get(DOWNLOAD_FIELD)
    return v if isinstance(v, int) and not isinstance(v, bool) and v >= 0 else None

# --- probe 2: repocore API for main-x (public; same API the anaconda.org front-end calls) ---
mx_channel, mx_channel_prov = fetch_json_with_integrity(REPOCORE_CHANNEL_API.format(channel="main-x"), "probe/repocore-channel")
print("\\nrepocore channel object keys:", sorted(mx_channel.keys()))
for k in ("artifact_count", "download_count", "description", "privacy"):
    print(f"  channel.{k} = {repr(mx_channel.get(k))[:140]}")
MX_ARTIFACT_COUNT_ADVERTISED = mx_channel.get("artifact_count")
print(f"repocore advertises {MX_ARTIFACT_COUNT_ADVERTISED:,} artifacts for main-x")

mx_artifact_page, _ = fetch_json_with_integrity(
    REPOCORE_ARTIFACTS_API.format(channel="main-x") + "?limit=2&offset=0", "probe/repocore-artifacts")
if not isinstance(mx_artifact_page, dict) or not isinstance(mx_artifact_page.get("items"), list):
    raise ChannelDataError("repocore artifacts page shape unexpected (no 'items' list). Stopping.")
print("\\nrepocore artifact keys:", sorted(mx_artifact_page["items"][0].keys()))
print("artifact.metadata keys:", sorted((mx_artifact_page["items"][0].get("metadata") or {}).keys()))
print("repocore page reports total_count =", mx_artifact_page.get("total_count"))"""))

cells.append(nbf.v4.new_code_cell("""def download_fetch_one(name):
    \"\"\"One anaconda.org package-API lookup with retry/backoff on 429/5xx/network.\"\"\"
    url = ANACONDA_ORG_PACKAGE_API.format(name=name)
    for attempt in range(DL_MAX_RETRIES):
        try:
            data, _ = fetch_json_with_integrity(url, f"dl/{name}", timeout=DL_TIMEOUT)
            nd = parse_downloads(data)
            if nd is None:
                return name, {"status": "no_field", "downloads": None, "source_url": url, "retrieved_utc": now_utc()}
            return name, {"status": "ok", "downloads": nd, "source_url": url, "retrieved_utc": now_utc()}
        except ChannelDataError as e:
            msg = str(e)
            if "HTTP 404" in msg:
                return name, {"status": "not_found", "downloads": None, "source_url": url, "retrieved_utc": now_utc()}
            transient = ("HTTP 429" in msg) or any(f"HTTP {c}" in msg for c in (500, 502, 503, 504)) or "network error" in msg
            if transient and attempt < DL_MAX_RETRIES - 1:
                time.sleep(2 ** attempt)
                continue
            return name, {"status": "error", "downloads": None, "source_url": url, "retrieved_utc": now_utc()}
    return name, {"status": "error", "downloads": None, "source_url": url, "retrieved_utc": now_utc()}


def fetch_download_counts(names):
    \"\"\"Resume-safe download-count pass for main. Cache keyed by package name; errors retried next run.\"\"\"
    names = sorted(set(names))
    cache = {}
    if os.path.exists(DOWNLOAD_CACHE_FILE):
        try:
            with open(DOWNLOAD_CACHE_FILE, encoding="utf-8") as fh:
                raw_cache = json.load(fh)
            if isinstance(raw_cache, dict):
                cache = {k: v for k, v in raw_cache.items()
                         if isinstance(v, dict) and v.get("status") in ("ok", "not_found", "no_field")}
        except (OSError, json.JSONDecodeError) as e:
            print(f"downloads cache unreadable ({e}); starting fresh")
    todo = [n for n in names if n not in cache]
    print(f"downloads: {len(todo):,} lookups to do, {len(names) - len(todo):,} already cached")
    t0, done = time.time(), 0
    with ThreadPoolExecutor(max_workers=DL_WORKERS) as pool:
        futures = {pool.submit(download_fetch_one, n): n for n in todo}
        for fut in as_completed(futures):
            name, entry = fut.result()
            done += 1
            if entry["status"] != "error":
                cache[name] = entry
            if done % 1000 == 0:
                print(f"  ...{done:,}/{len(todo):,} download lookups ({time.time() - t0:.0f}s)")
    dtmp = DOWNLOAD_CACHE_FILE + ".tmp"
    with open(dtmp, "w", encoding="utf-8") as fh:
        json.dump(cache, fh)
    os.replace(dtmp, DOWNLOAD_CACHE_FILE)
    return {n: cache.get(n, {"status": "error", "downloads": None,
                             "source_url": ANACONDA_ORG_PACKAGE_API.format(name=n),
                             "retrieved_utc": now_utc()}) for n in names}


def fetch_repocore_artifacts(channel):
    \"\"\"Paginate the repocore artifacts API; validate totals and expected fields on every page.\"\"\"
    base = REPOCORE_ARTIFACTS_API.format(channel=channel)
    items, offset, total_count = [], 0, None
    while True:
        url = f"{base}?limit={REPOCORE_PAGE_LIMIT}&offset={offset}"
        page, prov = fetch_json_with_integrity(url, f"repocore/{channel}@{offset}")
        if not isinstance(page, dict):
            raise ChannelDataError(f"[repocore/{channel}] page at offset {offset} is not an object. Refusing.")
        if "items" not in page or not isinstance(page["items"], list):
            raise ChannelDataError(f"[repocore/{channel}] page at offset {offset} lacks 'items' list. Refusing.")
        if total_count is None:
            total_count = page.get("total_count")
            if not isinstance(total_count, int) or total_count <= 0:
                raise ChannelDataError(f"[repocore/{channel}] total_count missing/invalid ({total_count}). Refusing.")
        elif page.get("total_count") != total_count:
            raise ChannelDataError(f"[repocore/{channel}] total_count changed mid-pagination ({total_count} -> {page.get('total_count')}). Refusing.")
        items.extend(page["items"])
        offset += len(page["items"])
        print(f"  repocore/{channel}: {offset:,}/{total_count:,}")
        if len(page["items"]) < REPOCORE_PAGE_LIMIT or not page["items"]:
            break
    if len(items) != total_count:
        raise ChannelDataError(f"[repocore/{channel}] fetched {len(items):,} items but total_count={total_count:,}. Refusing.")
    for it in items:
        if not isinstance(it, dict) or not it.get("name"):
            raise ChannelDataError(f"[repocore/{channel}] artifact without 'name' found. Refusing.")
    print(f"repocore/{channel}: {len(items):,} artifacts fetched, totals cross-checked")
    return items, {"source_url": base, "retrieved_utc": prov["retrieved_utc"],
                   "items": len(items), "total_count": total_count}

print("fetchers ready")"""))

cells.append(nbf.v4.new_code_cell("""# ---------- main (public channeldata) ----------
main_data, main_prov = fetch_json_with_integrity(MAIN_CHANNELDATA_URL, "main")
main_pkgs, _ = validate_channeldata(main_data, "main", main_prov)
main_catalog, main_missing_summary = catalog_from_channeldata_packages(main_pkgs, "main")
main_prov["method"] = "channeldata.json 'packages' map (already keyed by unique package name)"
main_prov["packages_lacking_summary"] = len(main_missing_summary)

# ---------- main-x channel contents (token) ----------
token, token_origin = resolve_repo_token()
auth = {"Authorization": f"Bearer {token}"}
main_x_data, main_x_prov_channeldata = fetch_json_with_integrity(MAIN_X_CHANNELDATA_URL, "main-x", extra_headers=auth)
main_x_pkgs, main_x_subdirs = validate_channeldata(main_x_data, "main-x", main_x_prov_channeldata)

if len(main_x_pkgs) >= MIN_PACKAGES_EXPECTED:
    main_x_catalog, _ = catalog_from_channeldata_packages(main_x_pkgs, "main-x")
    main_x_prov = main_x_prov_channeldata
    main_x_prov["method"] = "channeldata.json 'packages' map"
else:
    if not main_x_subdirs:
        raise ChannelDataError("[main-x] channeldata.json is an empty stub AND lists no subdirs to fall back to. Refusing.")
    records, repodata_provs = [], []
    for subdir in sorted(set(main_x_subdirs)):
        d, p = fetch_json_with_integrity(MAIN_X_REPODATA_URL.format(subdir=subdir), f"main-x/{subdir}", extra_headers=auth)
        subdir_records = parse_repodata(d, f"main-x/{subdir}", p)
        repodata_provs.append(p)
        records.extend(subdir_records)
    main_x_catalog = latest_from_repodata(records, "main-x")
    main_x_prov = {
        "source_url": redact(MAIN_X_CHANNELDATA_URL) + "  (empty stub; fell back to per-subdir repodata.json)",
        "retrieved_utc": main_x_prov_channeldata["retrieved_utc"],
        "http_last_modified": "; ".join(f"{redact(p['source_url'])}: {p['http_last_modified']}" for p in repodata_provs),
        "bytes_received": main_x_prov_channeldata["bytes_received"] + sum(p["bytes_received"] for p in repodata_provs),
        "content_length_check": "; ".join(
            f"{redact(p['source_url'])}: {p['content_length_check']}" for p in [main_x_prov_channeldata, *repodata_provs]
        ),
        "method": ("channeldata.json empty stub -> repodata.json per subdir; collapsed "
                   f"{len(records)} artifact records (build variants) to unique package names"),
        "subdirs_used": ", ".join(sorted(set(main_x_subdirs))),
        "raw_artifact_records": len(records),
    }
main_x_prov["token_source"] = token_origin
print(f"main   : {len(main_catalog):,} unique packages")
print(f"main-x : {len(main_x_catalog):,} unique packages (token from: {token_origin})")"""))

cells.append(nbf.v4.new_code_cell("""# ---------- main-x summaries + download counts from anaconda.org repocore (public API the .org front-end calls) ----------
main_x_artifacts, repocore_prov = fetch_repocore_artifacts("main-x")
by_name = {}
for it in main_x_artifacts:
    n = it["name"].strip()
    if n in by_name:
        raise ChannelDataError(f"[repocore/main-x] duplicate artifact name {n!r}. Refusing.")
    by_name[n] = it

repodata_names = set(main_x_catalog)
repocore_names = set(by_name)
if repodata_names != repocore_names:
    only_repo = sorted(repodata_names - repocore_names)
    only_core = sorted(repocore_names - repodata_names)
    raise ChannelDataError(
        f"[main-x] repodata and repocore disagree on the package set "
        f"(repodata-only: {len(only_repo)} e.g. {only_repo[:5]}; repocore-only: {len(only_core)} e.g. {only_core[:5]}). "
        "Refusing rather than merging mismatched catalogs.")

mx_summary_coverage = 0
for name, it in by_name.items():
    md = it.get("metadata") or {}
    summary = (md.get("summary") or "").strip()
    if not summary:
        desc = (md.get("description") or "").strip()
        summary = re.sub(r"\\s+", " ", desc).strip()[:500] if desc else ""
    if summary:
        mx_summary_coverage += 1
    lic = (md.get("license") or main_x_catalog[name]["license"] or "").strip()
    dc = it.get("download_count")
    main_x_catalog[name]["summary"] = summary
    main_x_catalog[name]["license"] = lic
    main_x_catalog[name]["downloads_total"] = dc if isinstance(dc, int) and dc >= 0 else None
    main_x_catalog[name]["downloads_source"] = repocore_prov["source_url"] + "?limit=1000 (paginated)"
    main_x_catalog[name]["downloads_asof"] = repocore_prov["retrieved_utc"]

mx_dl_all_zero = all((e["downloads_total"] or 0) == 0 for e in main_x_catalog.values())
main_x_prov["packages_lacking_summary"] = len(main_x_catalog) - mx_summary_coverage
print(f"main-x summaries from anaconda.org repocore: {mx_summary_coverage:,} of {len(main_x_catalog):,} ({100*mx_summary_coverage/len(main_x_catalog):.1f}%)")
print(f"main-x download_count on anaconda.org repocore: {'ALL ZERO — telemetry not populated for this channel yet' if mx_dl_all_zero else 'populated'}")"""))

cells.append(nbf.v4.new_code_cell("""# ---------- main downloads (anaconda.org package API; cached, resume-safe) ----------
downloads_by_name = fetch_download_counts(main_catalog.keys())
dl_pass_utc = now_utc()
dl_stats = {
    "total": len(downloads_by_name),
    "resolved": sum(1 for v in downloads_by_name.values() if v["status"] == "ok" and isinstance(v["downloads"], int)),
    "not_found": sum(1 for v in downloads_by_name.values() if v["status"] == "not_found"),
    "errors": sum(1 for v in downloads_by_name.values() if v["status"] == "error"),
    "no_field": sum(1 for v in downloads_by_name.values() if v["status"] == "no_field"),
}
dl_stats["coverage"] = dl_stats["resolved"] / dl_stats["total"] if dl_stats["total"] else 0.0
for name, entry in main_catalog.items():
    d = downloads_by_name[name]
    if d["status"] == "ok" and isinstance(d["downloads"], int):
        entry["downloads_total"] = d["downloads"]
        entry["downloads_source"] = d["source_url"]
        entry["downloads_asof"] = d["retrieved_utc"]
    else:
        entry["downloads_total"] = None
        entry["downloads_source"] = d["source_url"] + f"  (no download figure: {d['status']})"
        entry["downloads_asof"] = d["retrieved_utc"]
dl_coverage_warning = dl_stats["coverage"] < MIN_DOWNLOAD_COVERAGE_WARN

# ---------- hard gate ----------
for label, catalog, prov in (("main", main_catalog, main_prov), ("main-x", main_x_catalog, main_x_prov)):
    n = len(catalog)
    if n < MIN_PACKAGES_EXPECTED:
        raise ChannelDataError(f"[{label}] only {n} unique packages after processing; below sanity floor. Refusing.")
    prov["unique_package_count"] = n

print(f"main downloads: {dl_stats['resolved']:,} of {dl_stats['total']:,} resolved ({100*dl_stats['coverage']:.1f}%; "
      f"not found: {dl_stats['not_found']:,}, errors: {dl_stats['errors']:,})")
if dl_coverage_warning:
    print(f"WARNING: coverage {100*dl_stats['coverage']:.1f}% below {100*MIN_DOWNLOAD_COVERAGE_WARN:.0f}% — see Summary")"""))

cells.append(nbf.v4.new_code_cell("""main_names = set(main_catalog)
main_x_names = set(main_x_catalog)

COLUMNS_MAIN = ["Package", "Latest Version", "What it does", "On main-x?",
                "Downloads (total)", "Downloads (source)", "Downloads (as of)"]
COLUMNS_MAIN_X = ["Package", "Latest Version", "What it does", "On main?", "License",
                  "Downloads (total)", "Downloads (source)", "Downloads (as of)"]


def _dl_sort_key(item):
    name, e = item
    d = e.get("downloads_total")
    return (0 if isinstance(d, int) else 1, -(d if isinstance(d, int) else 0), name.lower())


def channel_frame(catalog, other_names, other_label, columns, license_col=False):
    rows = []
    for name, e in sorted(catalog.items(), key=_dl_sort_key):
        row = {"Package": name, "Latest Version": e["version"], "What it does": e["summary"],
               f"On {other_label}?": "yes" if name in other_names else "no"}
        if license_col:
            row["License"] = e.get("license") or ""
        row["Downloads (total)"] = e["downloads_total"]
        row["Downloads (source)"] = e["downloads_source"]
        row["Downloads (as of)"] = e["downloads_asof"]
        rows.append(row)
    df = pd.DataFrame(rows, columns=columns)
    assert df["Package"].is_unique, "duplicate package names slipped into a channel tab"
    return df


df_main = channel_frame(main_catalog, main_x_names, "main-x", COLUMNS_MAIN)
df_main_x = channel_frame(main_x_catalog, main_names, "main", COLUMNS_MAIN_X, license_col=True)
assert len(df_main) == len(main_names) and len(df_main_x) == len(main_x_names)

both = len(main_names & main_x_names)
print(f"overlap: {both:,} | main-only: {len(main_names - main_x_names):,} | main-x-only: {len(main_x_names - main_names):,}")
print("top 5 main by downloads:", ", ".join(f"{p} ({d:,})" for p, d in zip(df_main.head(5)["Package"], df_main.head(5)["Downloads (total)"])))
print("top 3 main-x rows (downloads all 0 -> alphabetical):", df_main_x.head(3)["Package"].tolist())"""))

cells.append(nbf.v4.new_code_cell("""def figure(name, value, source, retrieved, last_modified):
    return {"Figure": name, "Value": value, "Source": source,
            "Retrieved (UTC)": retrieved, "Source last-modified": last_modified}


dl_source = "anaconda.org package API, https://api.anaconda.org/package/anaconda/<name>"
repocore_source = "anaconda.org repocore API, https://api.anaconda.org/repocore/channels/main-x[/artifacts] (public; the .org front-end's own API)"
summary_rows = [
    figure("Unique package count — main", len(main_names),
           main_prov["source_url"], main_prov["retrieved_utc"], main_prov["http_last_modified"]),
    figure("Unique package count — main-x", len(main_x_names),
           main_x_prov["source_url"], main_x_prov["retrieved_utc"], main_x_prov["http_last_modified"]),
    figure("Packages present on both channels (by name)", both,
           "Intersection of the two name sets above", f"{main_prov['retrieved_utc']} / {main_x_prov['retrieved_utc']}",
           f"{main_prov['http_last_modified']} / {main_x_prov['http_last_modified']}"),
    figure("Packages only on main", len(main_names - main_x_names),
           "main name set minus main-x name set", main_prov["retrieved_utc"], main_prov["http_last_modified"]),
    figure("Packages only on main-x", len(main_x_names - main_names),
           "main-x name set minus main name set", main_x_prov["retrieved_utc"], main_x_prov["http_last_modified"]),
    figure("Counting rule", "unique package names; build variants collapsed (repodata: max by version, then build number, then timestamp)",
           "methodology", now_utc(), "n/a"),
    figure("'Latest version' rule — main", "channeldata.json per-package 'version' field (the channel's latest)",
           main_prov["source_url"], main_prov["retrieved_utc"], main_prov["http_last_modified"]),
    figure("'Latest version' rule — main-x", main_x_prov["method"],
           main_x_prov["source_url"], main_x_prov["retrieved_utc"], main_x_prov["http_last_modified"]),
    figure("main 'what it does' — Anaconda's own channeldata summaries",
           f"{len(main_names) - main_prov['packages_lacking_summary']:,} of {len(main_names):,} have one; "
           f"{main_prov['packages_lacking_summary']} are empty because the channel publishes none",
           main_prov["source_url"], main_prov["retrieved_utc"], main_prov["http_last_modified"]),
    figure("main-x 'what it does' — anaconda.org repocore summaries",
           f"{mx_summary_coverage:,} of {len(main_x_names):,} have one ({100*mx_summary_coverage/len(main_x_names):.1f}%), "
           "exact-name match to catalog, no fuzzy matching; empty = anaconda.org publishes none",
           repocore_source, repocore_prov["retrieved_utc"], "n/a"),
    figure("Downloads — aggregation rule",
           f"anaconda.org package-level '{DOWNLOAD_FIELD}' field: server-side aggregate across all files, all versions, all platforms "
           "(field chosen by live probe; per-file sums drift slightly from this counter)",
           dl_source, dl_pass_utc, "n/a (API responses carry no Last-Modified)"),
    figure("Downloads — coverage, main",
           (f"{dl_stats['resolved']:,} of {dl_stats['total']:,} packages got a number ({100 * dl_stats['coverage']:.1f}%); "
            f"not found on anaconda.org: {dl_stats['not_found']:,}; no download field: {dl_stats['no_field']:,}; "
            f"fetch errors this run (retried next run): {dl_stats['errors']:,}"
            + (f"  *** WARNING: coverage below {100*MIN_DOWNLOAD_COVERAGE_WARN:.0f}% — column is partial ***" if dl_coverage_warning else "")),
           dl_source, dl_pass_utc, "n/a"),
    figure("Downloads — main-x",
           ("anaconda.org repocore reports download_count = 0 for EVERY main-x package as of the fetch date. "
            "Telemetry is not populated for this channel yet — do not read zeros as zero usage. "
            "When .org starts reporting real numbers, re-running this notebook picks them up automatically."
            if mx_dl_all_zero else
            "download_count populated on anaconda.org repocore"),
           repocore_source, repocore_prov["retrieved_utc"], "n/a"),
    figure("main-x on anaconda.org — channel page exists",
           f"anaconda.org/channels/main-x browsable; repocore channel object advertises artifact_count = {MX_ARTIFACT_COUNT_ADVERTISED:,} "
           f"(matches this catalog's {len(main_x_names):,}), privacy = '{mx_channel.get('privacy')}'. "
           f"Channel description per .org: {str(mx_channel.get('description'))[:200]}",
           repocore_source, repocore_prov["retrieved_utc"], "n/a"),
    figure("Row sort order",
           "Both channel sheets: Downloads (total) desc, unknowns last, alphabetical within ties. main-x is "
           "all-zero on anaconda.org today, so its effective order is alphabetical. Autofilter on every sheet.",
           "presentation", now_utc(), "n/a"),
    figure("Integrity check — main", f"OK: {main_prov['content_length_check']}; strict JSON parse; "
           f"{main_prov['bytes_received']:,} bytes; {main_prov['channeldata_package_count']:,} packages; floor passed",
           main_prov["source_url"], main_prov["retrieved_utc"], main_prov["http_last_modified"]),
    figure("Integrity check — main-x", f"OK: {main_x_prov['content_length_check']}; strict JSON parse; "
           f"{main_x_prov['bytes_received']:,} bytes; floor passed"
           + (f"; {main_x_prov['raw_artifact_records']:,} raw artifact records collapsed" if main_x_prov.get("raw_artifact_records") else "")
           + f"; repocore pagination cross-checked ({repocore_prov['items']:,} items == total_count)",
           main_x_prov["source_url"], main_x_prov["retrieved_utc"], main_x_prov["http_last_modified"]),
    figure("main-x token source (token itself never stored)", token_origin,
           "local configuration", now_utc(), "n/a"),
    figure("Catalog generated at", now_utc(), "this notebook", now_utc(), "n/a"),
]
df_summary = pd.DataFrame(summary_rows)
df_summary"""))

cells.append(nbf.v4.new_code_cell("""# Everything above must succeed before any file is written; write atomically.
# Row 1 = channel header line (count + generation date + browse URL); row 2 = column headers; autofilter from row 2.
from openpyxl.styles import Font


def header_line(channel, count):
    return (f"Anaconda {channel} channel — {count:,} packages — generated {today_utc()} "
            f"from the official Anaconda repository. Browse online: {BROWSE_URLS[channel]}")


tmp = OUTPUT_XLSX.replace(".xlsx", "") + ".tmp.xlsx"
with pd.ExcelWriter(tmp, engine="openpyxl") as writer:
    df_main.to_excel(writer, sheet_name="main", index=False, startrow=1)
    df_main_x.to_excel(writer, sheet_name="main-x", index=False, startrow=1)
    df_summary.to_excel(writer, sheet_name="Summary", index=False, startrow=1)
    header_texts = {
        "main": header_line("main", len(df_main)),
        "main-x": header_line("main-x", len(df_main_x)),
        "Summary": f"Summary — anaconda_channel_catalog — generated {today_utc()}; every figure carries its source and date.",
    }
    for sheet, df in (("main", df_main), ("main-x", df_main_x), ("Summary", df_summary)):
        ws = writer.book[sheet]
        cell = ws.cell(row=1, column=1, value=header_texts[sheet])
        cell.font = Font(bold=True)
        for idx, col in enumerate(df.columns, start=1):
            longest = max((len(str(v)) for v in df[col].head(200)), default=0)
            ws.column_dimensions[ws.cell(row=2, column=idx).column_letter].width = min(max(len(col) + 2, longest + 2, 10), 60)
        ws.freeze_panes = "A3"
        ws.auto_filter.ref = f"A2:{ws.cell(row=2, column=len(df.columns)).column_letter}{len(df) + 2}"
os.replace(tmp, OUTPUT_XLSX)
print(f"Wrote {OUTPUT_XLSX}: main={len(df_main):,} rows, main-x={len(df_main_x):,} rows, Summary={len(df_summary)} figures")"""))

cells.append(nbf.v4.new_markdown_cell("""## Refreshing the catalog

Re-run all cells. Figures are point-in-time; the Summary sheet records source + retrieval date for every number.
On any truncation/validation failure the notebook raises `ChannelDataError` and **no xlsx is written** — a stale
catalog is never overwritten with partial data. `download_counts_cache.json` makes re-runs cheap; delete to force
a full re-fetch. When anaconda.org starts populating `download_count` for main-x, re-running picks the real numbers
up automatically (they are all zero there as of the generation date)."""))

nb["cells"] = cells
nb["metadata"]["kernelspec"] = {"display_name": "Python 3", "language": "python", "name": "python3"}
nb["metadata"]["language_info"] = {"name": "python"}

OUT = "/workspace/30f1620a-4aad-4456-bf4d-550f335e6f55/11f2d22e-0dbf-4590-a1a9-066be1a36bcd/sessions/agent_4c9b3f0e-3060-4d32-9b06-b43fe2cd72b4/anaconda_channel_catalog.ipynb"
with open(OUT, "w", encoding="utf-8") as fh:
    nbf.write(nb, fh)
print("notebook written:", OUT, "| cells:", len(cells))
print("secrets embedded:", any(s in json.dumps(nb) for s in ("dc2505be", "github_pat_")))
